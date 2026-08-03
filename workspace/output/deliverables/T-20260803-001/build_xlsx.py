"""shiire_list_3000.csv → ローカル .xlsx（サービスアカウント不要の確実な納品物）。
タブ: サマリ・前提 / 仕入れ条件合致(原石) / 全リスト。原石は純利益降順。
"""
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path("/Users/yukinori/Claude Code/ai-company-amazon_buppan/workspace/output/deliverables/T-20260803-001")
CSV_PATH = OUT / "shiire_list_3000.csv"
SUMMARY_PATH = OUT / "shiire_summary.json"
XLSX = OUT / "仕入れせどり_売れ筋x仕入れ先リスト_4446件.xlsx"

COLS = [
    ("rank_no", "No"), ("asin", "ASIN"), ("amazon_url", "Amazonページ"),
    ("name", "商品名"), ("category", "カテゴリ"), ("jan", "JAN"),
    ("amazon_price", "Amazon価格"), ("buybox", "BuyBox価格"),
    ("sales_rank", "ランク"), ("offer_count", "出品者数"),
    ("monthly_sales", "推定月販"), ("size", "サイズ区分"),
    ("buy", "仕入値(最安)"), ("buy_source", "仕入先"),
    ("net", "純利益"), ("margin_pct", "利益率%"), ("roi_pct", "ROI%"),
    ("verdict", "判定"), ("gem", "原石"), ("durable", "持続"),
    ("supplier_hint", "仕入先メモ"),
]
HEADER = [h for _, h in COLS]
KEYS = [k for k, _ in COLS]
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(color="FFFFFF", bold=True)
GEM_FILL = PatternFill("solid", fgColor="E2EFDA")


def truthy(v):
    return str(v).strip().lower() in ("true", "1", "yes")


def num(v, d=-1):
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def numify(row):
    out = []
    for k in KEYS:
        v = row.get(k, "")
        if k in ("amazon_price", "buybox", "sales_rank", "offer_count",
                 "monthly_sales", "buy", "net", "margin_pct", "roi_pct", "rank_no"):
            fv = num(v, None)
            out.append(fv if fv is not None else v)
        else:
            out.append(v)
    return out


def style_header(ws):
    for c in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER))}1"


def widths(ws):
    w = {"C": 34, "D": 46, "E": 12, "F": 15, "U": 26}
    for col, val in w.items():
        ws.column_dimensions[col].width = val


def main():
    rows = list(csv.DictReader(open(CSV_PATH, encoding="utf-8")))
    s = json.loads(SUMMARY_PATH.read_text(encoding="utf-8")) if SUMMARY_PATH.exists() else {}
    gems = sorted([r for r in rows if truthy(r.get("gem"))],
                  key=lambda r: num(r.get("net")), reverse=True)

    wb = Workbook()

    # 1) サマリ
    ws = wb.active
    ws.title = "サマリ・前提"
    n_jan = sum(1 for r in rows if r.get("jan"))
    n_match = sum(1 for r in rows if num(r.get("buy")) > 0)
    n_gem = len(gems)
    n_dur = sum(1 for r in rows if truthy(r.get("durable")))
    info = [
        ["仕入れせどり｜Amazon売れ筋 × 仕入れ先リスト", ""],
        ["チケット", "T-20260803-001"],
        ["作成", s.get("finished_at", "")],
        ["", ""],
        ["【手法】", s.get("method", "")],
        ["Finder条件", s.get("finder_filter", "")],
        ["仕入れ条件(判定)", s.get("criteria", "")],
        ["", ""],
        ["総リスト件数", len(rows)],
        ["JAN取得済", n_jan],
        ["仕入先マッチ(価格取得)", n_match],
        ["仕入れ条件合致=原石(tier1)", n_gem],
        ["うち持続(tier2)", n_dur],
        ["", ""],
        ["【注意】実購入は社長承認必須(§4.1金銭)。本表は候補リストまで。", ""],
        ["各候補はAmazon実ページ×仕入先ページで同一商品かを購入前に必ず目視確認。", ""],
        ["仕入値は税込最安の自動逆引き。入数/型番差の誤突合が残り得る点に留意。", ""],
    ]
    for r in info:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80

    # 2) 原石
    ws2 = wb.create_sheet("仕入れ条件合致(原石)")
    ws2.append(HEADER)
    for r in gems:
        ws2.append(numify(r))
    style_header(ws2)
    widths(ws2)

    # 3) 全リスト
    ws3 = wb.create_sheet("全リスト")
    ws3.append(HEADER)
    gem_idx = KEYS.index("gem")
    for r in rows:
        ws3.append(numify(r))
        if truthy(r.get("gem")):
            for c in range(1, len(HEADER) + 1):
                ws3.cell(row=ws3.max_row, column=c).fill = GEM_FILL
    style_header(ws3)
    widths(ws3)

    wb.save(XLSX)
    print("saved:", XLSX)
    print(f"rows={len(rows)} gems={n_gem} durable={n_dur} matched={n_match}")


if __name__ == "__main__":
    main()
