"""社長が共有済みの空Googleシートへ、エクセルの全タブ（全2,000行）をそのまま投入する。
各データタブに「Amazonページ」リンク列（https://www.amazon.co.jp/dp/{ASIN}）を追加。
前提: 対象シートを SA(sheets-writer@claude-session-sheets.iam.gserviceaccount.com) に
      「編集者」で共有済みであること（社長の1回だけの操作）。
実行: python3 populate_research_gsheet.py
"""
import sys
from pathlib import Path

import gspread
from openpyxl import load_workbook

CRED = "/Users/yukinori/.config/claude-session-sheets/credentials.json"
XLSX = Path(__file__).with_name("原石密度_リサーチデータ_2000件.xlsx")
# Drive MCP で作成済みの空シート（社長所有）。共有されたらここへ書き込む。
SHEET_ID = "1AAWGWaj1r6mP3Up297KxOq7F24faKYEYFlcdJlrBT_c"
TAB_ORDER = ["サマリ・前提", "買い候補ショートリスト", "カテゴリ別サマリ",
             "クリーン原石(誤突合除去)", "全データ"]


def tab_rows(ws):
    rows = [[("" if c is None else c) for c in r] for r in ws.iter_rows(values_only=True)]
    if not rows:
        return rows
    header = [str(h) for h in rows[0]]
    ai = next((i for i, h in enumerate(header) if h.lower() == "asin"), None)
    if ai is None:
        return rows
    out = []
    for ri, row in enumerate(rows):
        row = list(row) + [""] * (len(header) - len(row))
        if ri == 0:
            row.insert(ai + 1, "Amazonページ")
        else:
            a = str(row[ai]).strip()
            row.insert(ai + 1, f"https://www.amazon.co.jp/dp/{a}" if a else "")
        out.append(row)
    return out


def main():
    gc = gspread.service_account(filename=CRED)
    try:
        sh = gc.open_by_key(SHEET_ID)
        sh.sheet1  # アクセス確認（未共有ならここで例外）
    except Exception as e:
        print("シートを開けません（SAへの共有がまだの可能性）:", type(e).__name__)
        print("→ 対象シートを sheets-writer@claude-session-sheets.iam.gserviceaccount.com に『編集者』で共有してください。")
        sys.exit(1)

    wb = load_workbook(XLSX)
    order = [t for t in TAB_ORDER if t in wb.sheetnames] + [t for t in wb.sheetnames if t not in TAB_ORDER]
    first = True
    default_ws = sh.sheet1
    for tab in order:
        rows = tab_rows(wb[tab])
        ncols = max((len(r) for r in rows), default=1)
        rows = [r + [""] * (ncols - len(r)) for r in rows]
        if first:
            ws = default_ws
            ws.update_title(tab)
            ws.resize(rows=max(len(rows), 1), cols=max(ncols, 1))
            first = False
        else:
            ws = sh.add_worksheet(title=tab, rows=max(len(rows), 1), cols=max(ncols, 1))
        sh.values_update(f"'{tab}'!A1", params={"valueInputOption": "RAW"}, body={"values": rows})
        print(f"  wrote {tab}: {len(rows)}行 x {ncols}列")
    print("完了:", sh.url)


if __name__ == "__main__":
    main()
