"""リサーチデータ(density_v2)を新規スプレッドシート(.xlsx)に表形式でまとめる。

タブ構成:
  1. 全データ        : スキャン全行（Amazon条件抽出→JAN逆引き→利益/流動性判定）
  2. 原石候補        : gem=True の行＋Buy Box確認結果（社長が実際に見る棚）
  3. カテゴリ別サマリ  : カテゴリごとの突合率・原石率
  4. サマリ/前提     : 手法・抽出条件・全体歩留まり
"""
import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "T-20260521-005" / "code"))
from calc.profit import ProfitInput, calculate  # noqa: E402

SET_RE = re.compile(r"(\d+)\s*(本|個|枚|袋|セット|kg|ケース|パック|入)")
INB = {"small": 150, "standard_1": 250, "standard_2": 300, "large_1": 380, "large_2": 480}


def gem_ratio(r):
    try:
        return float(r["bb_confirmed"]) / float(r["buy"])
    except (ValueError, TypeError, ZeroDivisionError):
        return 0


def is_mismatch(r):
    """入数不一致・別モデル誤突合の疑い（差が非現実的に大きい）。"""
    rt = gem_ratio(r)
    if rt > 2.5:
        return True
    if rt > 1.8 and SET_RE.search(r.get("name", "")):
        return True
    return False


def net_at_buybox(r):
    """実売＝Buy Box価格で純利益を再計算。"""
    try:
        res = calculate(ProfitInput(
            wholesale_price=float(r["buy"]), amazon_price=float(r["bb_confirmed"]),
            category_key="default", size_key=r.get("size") or "standard_1",
            inbound_shipping=INB.get(r.get("size"), 250), other_costs=150))
        return round(res.net_profit), round(res.margin_rate * 100, 1)
    except (ValueError, TypeError):
        return "", ""

OUT = Path("/Users/yukinori/Claude Code/ai-company-amazon_buppan/workspace/output/deliverables/T-20260705-001")
XLSX = OUT / "原石密度_リサーチデータ_2000件.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="0B3D5C")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
GEM_FILL = PatternFill("solid", fgColor="FFF3D6")
LIQ_FILL = PatternFill("solid", fgColor="D9EAD3")

# CSV列 → 表示名
COLS = [
    ("category", "カテゴリ"), ("name", "商品名(先頭)"), ("asin", "ASIN"), ("jan", "JAN"),
    ("amazon_price", "Amazon価格"), ("buy", "仕入最安"), ("buy_source", "仕入先"),
    ("net", "純利益"), ("margin_pct", "利益率%"), ("verdict", "判定"),
    ("offer_count", "相乗り数"), ("monthly_sales", "月販(推定)"), ("sales_rank", "ランク"),
    ("buybox", "BuyBox(粗)"), ("size", "サイズ"),
    ("gem", "原石"), ("durable", "持続"), ("liquid", "流動性"),
]
CAT_JP = {
    "home_kitchen": "ホーム＆キッチン", "drugstore": "ドラッグストア", "diy_tools": "DIY・工具",
    "pet": "ペット", "toys": "おもちゃ", "hobby": "ホビー", "sports": "スポーツ",
    "office": "文房具", "baby": "ベビー", "appliances": "家電",
}


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}1"


def num(v):
    try:
        return float(v) if ("." in str(v)) else int(v)
    except (ValueError, TypeError):
        return v


def sheet_all(wb, rows):
    ws = wb.active
    ws.title = "全データ"
    ws.append([jp for _, jp in COLS])
    for r in rows:
        ws.append([num(r.get(k, "")) for k, _ in COLS])
    style_header(ws, len(COLS))
    # 原石/流動性を色付け
    gi = [k for k, _ in COLS].index("gem")
    li = [k for k, _ in COLS].index("liquid")
    for ri, r in enumerate(rows, start=2):
        if str(r.get("liquid")) == "True":
            for c in range(1, len(COLS) + 1):
                ws.cell(ri, c).fill = LIQ_FILL
        elif str(r.get("gem")) == "True":
            for c in range(1, len(COLS) + 1):
                ws.cell(ri, c).fill = GEM_FILL
    widths = [16, 34, 12, 14, 11, 10, 9, 9, 8, 8, 8, 10, 9, 11, 10, 7, 7, 8]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd


def sheet_gems(wb):
    p = OUT / "density_v2_gems_liquidity.csv"
    ws = wb.create_sheet("原石候補")
    if not p.exists():
        ws.append(["(原石候補なし)"]); return
    rows = list(csv.DictReader(open(p, encoding="utf-8")))
    cols = COLS + [("bb_confirmed", "BuyBox確認"), ("count_new_confirmed", "新品出品数確認")]
    ws.append([jp for _, jp in cols])
    for r in sorted(rows, key=lambda x: -(num(x.get("net")) or 0) if isinstance(num(x.get("net")), (int, float)) else 0):
        ws.append([num(r.get(k, "")) for k, _ in cols])
    style_header(ws, len(cols))
    for i, wd in enumerate([16, 34, 12, 14, 11, 10, 9, 9, 8, 8, 8, 10, 9, 11, 10, 7, 7, 8, 11, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd


def sheet_clean_gems(wb):
    """誤突合を除去したクリーンな流動性原石（社長が最初に見るべき棚）。"""
    p = OUT / "density_v2_gems_liquidity.csv"
    ws = wb.create_sheet("クリーン原石(誤突合除去)")
    if not p.exists():
        ws.append(["(データなし)"]); return 0, 0
    liq = [r for r in csv.DictReader(open(p, encoding="utf-8")) if r.get("liquid") == "True"]
    clean = [r for r in liq if not is_mismatch(r)]
    cols = [("category", "カテゴリ"), ("name", "商品名"), ("asin", "ASIN"), ("jan", "JAN"),
            ("buy", "仕入最安"), ("buy_source", "仕入先"), ("bb_confirmed", "BuyBox実価格"),
            ("_ratio", "差(倍)"), ("_net_bb", "純利益(BB基準)"), ("_mg_bb", "利益率%(BB)"),
            ("offer_count", "相乗り数"), ("monthly_sales", "月販(推定)"), ("sales_rank", "ランク")]
    ws.append([jp for _, jp in cols])
    def netkey(r):
        n, _ = net_at_buybox(r)
        return n if isinstance(n, (int, float)) else 0
    for r in sorted(clean, key=lambda r: -netkey(r)):
        n, mg = net_at_buybox(r)
        vals = []
        for k, _ in cols:
            if k == "_ratio":
                vals.append(round(gem_ratio(r), 2))
            elif k == "_net_bb":
                vals.append(n)
            elif k == "_mg_bb":
                vals.append(mg)
            else:
                vals.append(num(r.get(k, "")))
        ws.append(vals)
    style_header(ws, len(cols))
    for i, wd in enumerate([16, 34, 12, 14, 10, 8, 12, 8, 13, 11, 9, 10, 9], start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    return len(liq), len(clean)


def sheet_cat(wb, summ):
    ws = wb.create_sheet("カテゴリ別サマリ")
    ws.append(["カテゴリ", "詳細取得", "仕入突合", "原石", "突合率%", "原石率%(対突合)"])
    for ck, s in (summ.get("by_category") or {}).items():
        d, m, g = s.get("detail", 0), s.get("matched", 0), s.get("gem", 0)
        ws.append([CAT_JP.get(ck, ck), d, m, g,
                   round(m / d * 100, 1) if d else 0,
                   round(g / m * 100, 1) if m else 0])
    style_header(ws, 6)
    for i, wd in enumerate([18, 10, 10, 8, 10, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd


def sheet_summary(wb, summ):
    ws = wb.create_sheet("サマリ・前提")
    pairs = [
        ("測定手法", summ.get("method", "")),
        ("Amazon側 抽出条件", summ.get("finder_filter", "")),
        ("原石/持続/流動性の定義", summ.get("criteria", "")),
        ("", ""),
        ("条件合致ASIN(抽出)", summ.get("asin_found")),
        ("詳細取得", summ.get("detailed")),
        ("JANあり", summ.get("has_jan")),
        ("仕入先が付いた(突合)", summ.get("supplier_matched")),
        ("突合率%", summ.get("match_rate_pct")),
        ("", ""),
        ("原石 Tier1(率8%&利益300)", summ.get("gem_tier1")),
        ("持続 Tier2(+相乗り2-10&月販3+)", summ.get("durable_tier2")),
        ("流動性 Tier3(+BuyBox実在)【生】", summ.get("liquid_tier3")),
        ("流動性率%(対 突合)【生】", summ.get("liquid_rate_of_matched_pct")),
        ("", ""),
        ("★誤突合除去後 クリーン流動性原石", summ.get("_clean_liquid")),
        ("★クリーン原石率%(対 突合)", summ.get("_clean_rate_of_matched")),
        ("除去した誤突合(入数不一致/別モデル)", (summ.get("_liquid_raw") or 0) - (summ.get("_clean_liquid") or 0)),
        ("", ""),
        ("Keepa残トークン", summ.get("tokens_left")),
        ("重要注記", "生の流動性14.7%は入数不一致(消火器1本→19本セット等)・別モデル誤突合で水増し。差(BB÷仕入)>2.5倍を除去したクリーン値が実力。さらにクリーン原石の約半分はブランド出品制限品(初心者は相乗り不可)。月販はrank由来の粗推定。前回(Yahoo235件)との違い=母集団をAmazon側PoiPoi条件で選別済み。"),
    ]
    ws.append(["項目", "値"])
    for k, v in pairs:
        ws.append([k, v])
    style_header(ws, 2)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def main():
    rows = list(csv.DictReader(open(OUT / "density_v2_results.csv", encoding="utf-8")))
    summ = json.loads((OUT / "density_v2_summary.json").read_text(encoding="utf-8"))
    wb = Workbook()
    sheet_all(wb, rows)
    n_liq, n_clean = sheet_clean_gems(wb)
    sheet_gems(wb)
    sheet_cat(wb, summ)
    # 誤突合除去後の補正値をサマリに追記
    matched = summ.get("supplier_matched") or 1
    summ["_liquid_raw"] = n_liq
    summ["_clean_liquid"] = n_clean
    summ["_clean_rate_of_matched"] = round(n_clean / matched * 100, 1)
    sheet_summary(wb, summ)
    wb.move_sheet("サマリ・前提", -(len(wb.sheetnames) - 1))
    wb.save(XLSX)
    print(f"保存: {XLSX} ({len(rows)}行, タブ{len(wb.sheetnames)}, クリーン原石{n_clean}件)")


if __name__ == "__main__":
    main()
