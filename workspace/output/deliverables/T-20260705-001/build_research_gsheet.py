"""エクセル(原石密度_リサーチデータ_2000件.xlsx)の全タブをそのまま新規Googleスプレッドシートへ。
サービスアカウント(セッションシート用の既存鍵)で作成し、社長アカウントへ共有。
各データタブに「Amazonページ」リンク列（https://www.amazon.co.jp/dp/{ASIN}）を追加する。
"""
import sys
from pathlib import Path

import gspread
from openpyxl import load_workbook

CRED = "/Users/yukinori/.config/claude-session-sheets/credentials.json"
XLSX = Path("/Users/yukinori/Claude Code/ai-company-amazon_buppan/workspace/output/deliverables/T-20260705-001/原石密度_リサーチデータ_2000件.xlsx")
OWNER_EMAIL = "satoyukinori1018@gmail.com"
TITLE = "原石密度リサーチ_全データ_2000件（Amazonリンク付）"

# タブの並び（サマリ→アクション→データの順）
TAB_ORDER = ["サマリ・前提", "買い候補ショートリスト", "カテゴリ別サマリ",
             "クリーン原石(誤突合除去)", "全データ"]


def amazon_url(asin):
    asin = str(asin).strip()
    return f"https://www.amazon.co.jp/dp/{asin}" if asin else ""


def tab_rows(ws):
    """openpyxlワークシート→2次元リスト。asin列があれば直後にAmazonリンク列を挿入。"""
    rows = [[("" if c is None else c) for c in r] for r in ws.iter_rows(values_only=True)]
    if not rows:
        return rows
    header = [str(h) for h in rows[0]]
    asin_idx = next((i for i, h in enumerate(header) if h.lower() == "asin"), None)
    if asin_idx is None:
        return rows
    out = []
    for ri, row in enumerate(rows):
        row = list(row)
        # 列数をヘッダに合わせてパディング
        while len(row) < len(header):
            row.append("")
        if ri == 0:
            row.insert(asin_idx + 1, "Amazonページ")
        else:
            row.insert(asin_idx + 1, amazon_url(row[asin_idx]))
        out.append(row)
    return out


def main():
    gc = gspread.service_account(filename=CRED)
    sh = gc.create(TITLE)
    print("created:", sh.id)
    # 社長へ共有（編集可）
    sh.share(OWNER_EMAIL, perm_type="user", role="writer", notify=False)
    # リンクを知っていれば閲覧可にもしておく
    try:
        sh.share(None, perm_type="anyone", role="reader", with_link=True)
    except Exception as e:
        print("anyone-link share skipped:", e)

    wb = load_workbook(XLSX)
    order = [t for t in TAB_ORDER if t in wb.sheetnames] + [t for t in wb.sheetnames if t not in TAB_ORDER]
    first = True
    default_ws = sh.sheet1
    for tab in order:
        rows = tab_rows(wb[tab])
        ncols = max((len(r) for r in rows), default=1)
        # 全行を同じ列数にパディング
        rows = [r + [""] * (ncols - len(r)) for r in rows]
        if first:
            ws = default_ws
            ws.update_title(tab)
            ws.resize(rows=max(len(rows), 1), cols=max(ncols, 1))
            first = False
        else:
            ws = sh.add_worksheet(title=tab, rows=max(len(rows), 1), cols=max(ncols, 1))
        sh.values_update(
            f"'{tab}'!A1",
            params={"valueInputOption": "RAW"},
            body={"values": rows},
        )
        print(f"  wrote {tab}: {len(rows)}行 x {ncols}列")
    print("URL:", sh.url)


if __name__ == "__main__":
    main()
